import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import numbers


def create_excel_files(excel_path, lieferanten, output_dir='./files/bewertungsBoegen'):
    # Lade die Excel-Datei
    df = pd.read_excel(excel_path)

    # Erstelle das Ausgabeverzeichnis, falls es nicht existiert
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Hole alle einzigartigen Werte aus der Spalte "Zuständigkeit"
    zuständigkeiten = df['Zuständigkeit'].unique()

    for zuständigkeit in zuständigkeiten:
        # Filtere die Daten nach der aktuellen Zuständigkeit
        filtered_df = df[df['Zuständigkeit'] == zuständigkeit]

        # Erstelle eine neue Excel-Datei
        wb = Workbook()

        for lieferant in lieferanten:
            # Füge ein neues Blatt für jeden Lieferanten hinzu
            ws = wb.create_sheet(title=lieferant)

            # Schreibe die Spaltennamen
            ws.append(['Kriterien', 'Unterkriterien', 'Gewichtung', 'Bewertung'])

            # Schreibe die gefilterten Daten in das Blatt
            for index, row in filtered_df.iterrows():
                ws.append([row['Kriterien'], row['Unterkriterien'], row['Gewichtung'], 0])

            # Formatieren der Spalte "Gewichtung" als Prozent
            for cell in ws['C']:
                if cell.row == 1:
                    continue  # Überspringe die Kopfzeile
                cell.number_format = '0.00%'

        # Entferne das standardmäßig erstellte leere Blatt
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Speichere die neue Datei im angegebenen Ausgabeverzeichnis mit dem Namen der Zuständigkeit
        filename = os.path.join(output_dir, f"{zuständigkeit}.xlsx")
        wb.save(filename)

