import os
import datetime
import re
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def update_bewertungen(input_file, directory):
    # 1. Excel-Datei öffnen und Seitenamen extrahieren
    try:
        workbook = load_workbook(filename=input_file, read_only=True)
    except InvalidFileException:
        print(f"Error: Could not open {input_file}. Please check if it's a valid Excel file.")
        return

    sheet_names = workbook.sheetnames

    # 2. Seitenamen normalisieren (kleinbuchstaben, ohne sonderzeichen und leerzeichen)
    normalized_names = [re.sub(r'\W+', '', name.lower()) for name in sheet_names]

    # 3. Überprüfen und Seiten in vorhandene Dateien einfügen
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            for norm_name, original_name in zip(normalized_names, sheet_names):
                if norm_name in filename.lower():
                    # Seiten aus der Eingabedatei in die vorhandene Datei einfügen
                    try:
                        target_workbook = load_workbook(filename=os.path.join(directory, filename))
                    except InvalidFileException:
                        print(f"Error: Konnte nicht geöffnet werden  {os.path.join(directory, filename)}")
                        continue

                    # Aktuelles Datum im Format "yyyy-mm-dd" holen
                    current_date = datetime.datetime.now().strftime('%Y-%m-%d')

                    # Neue Seite mit dem aktuellen Datum als Name hinzufügen
                    new_sheet_name = current_date
                    new_sheet = target_workbook.create_sheet(title=new_sheet_name)

                    # Kopieren der Seite aus der Eingabedatei in die neue Seite
                    source_sheet = workbook[original_name]
                    for row in source_sheet.iter_rows(values_only=True):
                        new_sheet.append(row)

                    # Speichern der Änderungen
                    target_workbook.save(filename=os.path.join(directory, filename))
                    print(
                        f"Seite  '{original_name}'  wurde zu hinzugefügt '{os.path.join(directory, filename)}' als '{new_sheet_name}'.")
    print("Update process complete.")


# # Beispielaufruf:
# input_file = '../files/ausführendeBewertung.xlsx'
# directory = '../files/bewertungen'
# update_excel_pages(input_file, directory)
