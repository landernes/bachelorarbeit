import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def check_excel_files(directory):
    """
    Prüft alle Excel-Dateien in einem Verzeichnis darauf, ob in den Spalten "Bewertung" und "Unterkriterien"
    die Bedingung, dass die Bewertungen vorhanden sind, erfüllt ist

    Returns:
    bool: True, wenn alle Dateien die Bedingungen erfüllen, sonst False.
    """
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            file_path = os.path.join(directory, filename)
            xls = pd.ExcelFile(file_path)

            for sheet_name in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet_name)

                # Überprüfe, ob beide Spalten vorhanden sind
                if 'Bewertung' not in df.columns or 'Unterkriterien' not in df.columns:
                    return False

                # Überprüfe jede Zeile, in der 'Unterkriterien' nicht leer ist
                for index, row in df.iterrows():
                    if pd.notna(row['Unterkriterien']) and (pd.isna(row['Bewertung']) or row['Bewertung'] == 0):
                        return False

    return True



def clone_worksheet(input_file, directory):
    """
    Klont die erste Seite in der Eingabedatei für jeden einzigartigen Namenswert der Seiten
    aus den Excel-Dateien im Verzeichnis.
    """
    if not check_excel_files(directory):
        print("Keine Excel-Dateien im Verzeichnis gefunden.")
        return

    # Lade die Eingabedatei
    input_wb = openpyxl.load_workbook(input_file)
    input_ws = input_wb.active

    # Lösche alle Blätter außer dem ersten
    sheetnames = input_wb.sheetnames
    for sheetname in sheetnames[1:]:
        std = input_wb[sheetname]
        input_wb.remove(std)

    unique_sheet_names = set()

    # Durchsuche das Verzeichnis nach Excel-Dateien
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
            file_path = os.path.join(directory, filename)
            wb = openpyxl.load_workbook(file_path, read_only=True)
            unique_sheet_names.update(wb.sheetnames)

    # Klone das erste Blatt für jeden einzigartigen Blattnamen
    for sheet_name in unique_sheet_names:
        if sheet_name in input_wb.sheetnames:
            continue
        cloned_ws = input_wb.copy_worksheet(input_ws)
        cloned_ws.title = sheet_name

    # Speichere die Änderungen in der Eingabedatei
    input_wb.save(input_file)
    print(f"Die Seiten wurden erfolgreich in {input_file} geklont.")


def update_ratings(input_file, directory):
    # Prüfen, ob es Excel-Dateien im Verzeichnis gibt
    if not check_excel_files(directory):
        print("Keine Excel-Dateien im Verzeichnis gefunden.")
        return

    # Eingabedatei einlesen und erstes Blatt ignorieren
    input_wb = load_workbook(input_file)
    sheet_names = input_wb.sheetnames
    sheets_to_update = sheet_names[1:]  # Ignoriert das erste Blatt

    for sheet_name in sheets_to_update:
        df_input_sheet = pd.read_excel(input_file, sheet_name=sheet_name)
        input_sheet = input_wb[sheet_name]

        for file in os.listdir(directory):
            if file.endswith('.xlsx'):
                file_path = os.path.join(directory, file)
                xl = pd.ExcelFile(file_path)

                for xl_sheet_name in xl.sheet_names:
                    if xl_sheet_name == sheet_name:
                        xl_sheet = xl.parse(xl_sheet_name)

                        if 'Unterkriterien' in df_input_sheet.columns and 'Bewertung' in xl_sheet.columns:
                            for index, row in df_input_sheet.iterrows():
                                unterkriterium = row['Unterkriterien']
                                match = xl_sheet[xl_sheet['Unterkriterien'] == unterkriterium]

                                if not match.empty:
                                    bewertung = match.iloc[0]['Bewertung']
                                    # Spalte 'Bewertung' in der Eingabedatei aktualisieren
                                    input_sheet.cell(row=index + 2, column=df_input_sheet.columns.get_loc(
                                        'Bewertung') + 1).value = bewertung

    # Eingabedatei speichern
    input_wb.save(input_file)
    print("Die Eingabedatei wurde erfolgreich aktualisiert.")

"Speichern muss angepasst werden, sorgt für abstürzen da speichern nuter dialog sich öffnet"

def evaluate(input_file, directory):
    clone_worksheet(input_file, directory)
    update_ratings(input_file, directory)

# # Beispielaufruf der Funktion
# input_file = '../files/ausführendeBewertungTest.xlsx'
# directory = '../files/erwarteteBoegen'


