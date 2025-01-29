import xlwings as xw
import json
import openpyxl
from datetime import datetime


def calculate_and_save_excel(excel_file):
    app = xw.App(visible=False)
    wb = xw.Book(excel_file)
    wb.save()
    wb.close()
    app.quit()


def update_lieferanten_json(excel_file, json_file):
    # Berechnen und Speichern der Excel-Datei mit xlwings
    calculate_and_save_excel(excel_file)

    # Laden der JSON-Datei
    with open(json_file, 'r') as f:
        json_data = json.load(f)

    # Laden der Excel-Datei mit den berechneten Werten
    workbook = openpyxl.load_workbook(excel_file, data_only=True)

    # Aktuelles Datum
    current_date = datetime.now().strftime("%Y-%m-%d")

    # Durchlaufen aller Blätter in der Excel-Datei
    for sheet_name in workbook.sheetnames:
        # Überprüfen, ob der Blattname in den JSON-Daten enthalten ist
        for item in json_data:
            if item["Name"] == sheet_name:
                sheet = workbook[sheet_name]

                # Suchen der Zelle mit "Gesamtergebnis" und den Wert rechts davon
                for row in sheet.iter_rows(values_only=True):
                    for col_idx, cell_value in enumerate(row):
                        if cell_value == "Gesamtergebnis":
                            if col_idx + 1 < len(row):
                                value = row[col_idx + 1]
                                if isinstance(value, (int, float)):  # Überprüfen, ob der Wert eine Zahl ist
                                    value = round(value, 2)  # Auf 2 Dezimalstellen runden
                                item["Bewertungsscore"] = value
                                item["LetzteBewertung"] = current_date
                                break

    # Speichern der aktualisierten JSON-Daten
    with open(json_file, 'w') as f:
        json.dump(json_data, f, indent=4)


# # Beispielaufruf:
# input_file = '../files/ausführendeBewertung.xlsx'
# output_file = '../files/lieferanten.json'
# update_lieferanten_json(input_file, output_file)
